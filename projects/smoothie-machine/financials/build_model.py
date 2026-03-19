from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── helpers ────────────────────────────────────────────────────────────────────
BLUE  = "FF0000FF"
BLACK = "FF000000"
GREEN = "FF008000"
WHITE = "FFFFFFFF"
YELLOW_BG = "FFFFFF00"
DARK_BG   = "FF1F3864"
MID_BG    = "FF2E5090"
LIGHT_BG  = "FFD9E1F2"
ALT_BG    = "FFEEF2FB"

def font(bold=False, color=BLACK, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s = Side(style="medium", color="FF1F3864")
    return Border(bottom=s)

def euro(ws, cell, formula_or_value, is_input=False):
    c = ws[cell]
    c.value = formula_or_value
    c.number_format = u'#,##0.00\u20ac'
    c.font = font(color=BLUE if is_input else BLACK)
    c.alignment = align("right")

def pct(ws, cell, formula_or_value, is_input=False):
    c = ws[cell]
    c.value = formula_or_value
    c.number_format = "0.0%"
    c.font = font(color=BLUE if is_input else BLACK)
    c.alignment = align("right")

def num(ws, cell, formula_or_value, is_input=False):
    c = ws[cell]
    c.value = formula_or_value
    c.number_format = "#,##0"
    c.font = font(color=BLUE if is_input else BLACK)
    c.alignment = align("right")

def label(ws, cell, text, bold=False, size=11, color=BLACK, h="left"):
    c = ws[cell]
    c.value = text
    c.font = font(bold=bold, color=color, size=size)
    c.alignment = align(h)

def section_header(ws, cell, text, col_end, row=None):
    r = int(cell[1:]) if row is None else row
    col_start = cell[0]
    ws[cell] = text
    ws[cell].font = font(bold=True, color=WHITE, size=11)
    ws[cell].fill = fill(DARK_BG)
    ws[cell].alignment = align("left", "center")
    ws.merge_cells(f"{col_start}{r}:{col_end}{r}")

def sub_header(ws, cell, text, col_end, row=None):
    r = int(cell[1:]) if row is None else row
    col_start = cell[0]
    ws[cell] = text
    ws[cell].font = font(bold=True, color=WHITE, size=10)
    ws[cell].fill = fill(MID_BG)
    ws[cell].alignment = align("left", "center")
    ws.merge_cells(f"{col_start}{r}:{col_end}{r}")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Assumptions"
ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 28

# Title
ws1.merge_cells("A1:C1")
ws1["A1"] = "Smoothie Machine — Business Model Assumptions"
ws1["A1"].font = font(bold=True, size=14, color=WHITE)
ws1["A1"].fill = fill(DARK_BG)
ws1["A1"].alignment = align("center", "center")
ws1.row_dimensions[1].height = 28

ws1.merge_cells("A2:C2")
ws1["A2"] = "Blue = input  |  Black = calculated"
ws1["A2"].font = font(italic=True, size=9, color="FF555555")
ws1["A2"].alignment = align("center")

# ── CAPEX ─────────────────────────────────────────────────────────────────────
section_header(ws1, "A4", "CAPITAL EXPENDITURE (CAPEX)", "C")
ws1.row_dimensions[4].height = 20

rows_capex = [
    ("HM-160E machine (landed, DHL)", 4000, "€ — from Hommy quote 2026-03-12"),
    ("Sealing machine (tabletop)", 250, "€ — manual cup sealer"),
    ("Mini display fridge (clear door)", 350, "€ — holds pre-packed cups at venue"),
    ("Branding & signage", 400, "€ — first location"),
    ("Initial setup / miscellaneous", 150, "€ — tools, cleaning supplies"),
]
for i, (lbl, val, note) in enumerate(rows_capex, start=5):
    label(ws1, f"A{i}", lbl)
    euro(ws1, f"B{i}", val, is_input=True)
    label(ws1, f"C{i}", note, color="FF555555", size=9)
    ws1.row_dimensions[i].height = 16

total_capex_row = 5 + len(rows_capex)
label(ws1, f"A{total_capex_row}", "TOTAL CAPEX", bold=True)
euro(ws1, f"B{total_capex_row}", f"=SUM(B5:B{total_capex_row-1})")
ws1[f"A{total_capex_row}"].fill = fill(LIGHT_BG)
ws1[f"B{total_capex_row}"].fill = fill(LIGHT_BG)
ws1[f"B{total_capex_row}"].font = font(bold=True)
ws1.row_dimensions[total_capex_row].height = 18

# ── VARIABLE COSTS PER CUP ────────────────────────────────────────────────────
r = total_capex_row + 2
section_header(ws1, f"A{r}", "VARIABLE COSTS PER CUP", "C")
ws1.row_dimensions[r].height = 20
r += 1

var_rows_start = r
rows_var = [
    ("Frozen fruit mix (250g @ €3.20/kg)", 0.80, "€ — Makro Portugal, frozen blend"),
    ("Cup + lid (PLA, Vegware)", 0.31, "€ — 500 units @ €0.31/unit"),
    ("Label / packaging print", 0.05, "€ — custom sticker per cup"),
    ("Electricity per cup", 0.05, "€ — ~0.01 kWh × €0.22/kWh × blends"),
    ("Cleaning supplies (amortised)", 0.03, "€ — sanitiser, cloths"),
]
for lbl, val, note in rows_var:
    label(ws1, f"A{r}", lbl)
    euro(ws1, f"B{r}", val, is_input=True)
    label(ws1, f"C{r}", note, color="FF555555", size=9)
    ws1.row_dimensions[r].height = 16
    r += 1

var_rows_end = r - 1
label(ws1, f"A{r}", "TOTAL COGS per cup", bold=True)
euro(ws1, f"B{r}", f"=SUM(B{var_rows_start}:B{var_rows_end})")
ws1[f"A{r}"].fill = fill(LIGHT_BG)
ws1[f"B{r}"].fill = fill(LIGHT_BG)
ws1[f"B{r}"].font = font(bold=True)
COGS_CELL = f"B{r}"   # reference for other sheets
ws1.row_dimensions[r].height = 18
r += 2

# ── PRICING & FEES ────────────────────────────────────────────────────────────
section_header(ws1, f"A{r}", "PRICING & REVENUE SHARE", "C")
ws1.row_dimensions[r].height = 20
r += 1

label(ws1, f"A{r}", "Selling price per cup")
euro(ws1, f"B{r}", 4.00, is_input=True)
label(ws1, f"C{r}", "€ — target retail price", color="FF555555", size=9)
PRICE_CELL = f"B{r}"
ws1.row_dimensions[r].height = 16
r += 1

label(ws1, f"A{r}", "Venue revenue share")
pct(ws1, f"B{r}", 0.10, is_input=True)
label(ws1, f"C{r}", "% of revenue paid to venue owner", color="FF555555", size=9)
VENUE_PCT_CELL = f"B{r}"
ws1.row_dimensions[r].height = 16
r += 1

label(ws1, f"A{r}", "Venue fee per cup (calculated)", bold=False)
euro(ws1, f"B{r}", f"={PRICE_CELL}*{VENUE_PCT_CELL}")
ws1.row_dimensions[r].height = 16
VENUE_FEE_CELL = f"B{r}"
r += 2

# ── FIXED MONTHLY COSTS ───────────────────────────────────────────────────────
section_header(ws1, f"A{r}", "FIXED MONTHLY COSTS", "C")
ws1.row_dimensions[r].height = 20
r += 1

fixed_start = r
label(ws1, f"A{r}", "Machine depreciation (36 months)")
euro(ws1, f"B{r}", "=B5/36")
label(ws1, f"C{r}", "€ — straight-line over 3 years", color="FF555555", size=9)
ws1.row_dimensions[r].height = 16
r += 1

label(ws1, f"A{r}", "Sealing machine depreciation (36 mo)")
euro(ws1, f"B{r}", "=B6/36")
ws1.row_dimensions[r].height = 16
r += 1

label(ws1, f"A{r}", "Regulatory / permits (annual ÷ 12)")
euro(ws1, f"B{r}", 67, is_input=True)
label(ws1, f"C{r}", "€ — ASAE + hygiene cert amortised", color="FF555555", size=9)
ws1.row_dimensions[r].height = 16
r += 1

label(ws1, f"A{r}", "Insurance")
euro(ws1, f"B{r}", 30, is_input=True)
label(ws1, f"C{r}", "€ — estimate", color="FF555555", size=9)
ws1.row_dimensions[r].height = 16
r += 1

fixed_end = r - 1
label(ws1, f"A{r}", "TOTAL Fixed Costs / month", bold=True)
euro(ws1, f"B{r}", f"=SUM(B{fixed_start}:B{fixed_end})")
ws1[f"A{r}"].fill = fill(LIGHT_BG)
ws1[f"B{r}"].fill = fill(LIGHT_BG)
ws1[f"B{r}"].font = font(bold=True)
FIXED_CELL = f"B{r}"
ws1.row_dimensions[r].height = 18
r += 2

# ── TAXES ─────────────────────────────────────────────────────────────────────
section_header(ws1, f"A{r}", "TAXES", "C")
ws1.row_dimensions[r].height = 20
r += 1

label(ws1, f"A{r}", "IRC corporate tax rate")
pct(ws1, f"B{r}", 0.21, is_input=True)
label(ws1, f"C{r}", "% — Portugal Lda standard rate (2026)", color="FF555555", size=9)
TAX_CELL = f"B{r}"
ws1.row_dimensions[r].height = 16
r += 1

label(ws1, f"A{r}", "Derrama municipal surcharge")
pct(ws1, f"B{r}", 0.015, is_input=True)
label(ws1, f"C{r}", "% — max 1.5%, varies by municipality", color="FF555555", size=9)
DERRAMA_CELL = f"B{r}"
ws1.row_dimensions[r].height = 16
r += 1

label(ws1, f"A{r}", "Effective tax rate (IRC + derrama)", bold=False)
pct(ws1, f"B{r}", f"={TAX_CELL}+{DERRAMA_CELL}")
ws1[f"A{r}"].fill = fill(LIGHT_BG)
ws1[f"B{r}"].fill = fill(LIGHT_BG)
EFFECTIVE_TAX_CELL = f"B{r}"
ws1.row_dimensions[r].height = 16
r += 1

ws1.merge_cells(f"A{r}:C{r}")
ws1[f"A{r}"] = "Note: tax applies only on positive profit. Losses carry forward."
ws1[f"A{r}"].font = font(italic=True, size=9, color="FF555555")
ws1[f"A{r}"].alignment = align("left")
ws1.row_dimensions[r].height = 14

# Legend note at bottom
r += 2
ws1.merge_cells(f"A{r}:C{r}")
ws1[f"A{r}"] = "Blue cells are inputs — adjust freely. All other values update automatically."
ws1[f"A{r}"].font = font(italic=True, size=9, color="FF1F3864")
ws1[f"A{r}"].alignment = align("center")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — UNIT ECONOMICS
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Unit Economics")
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 22

ws2.merge_cells("A1:C1")
ws2["A1"] = "Unit Economics — Per Cup Sold"
ws2["A1"].font = font(bold=True, size=14, color=WHITE)
ws2["A1"].fill = fill(DARK_BG)
ws2["A1"].alignment = align("center", "center")
ws2.row_dimensions[1].height = 28

rows_ue = [
    ("Selling price", f"=Assumptions!{PRICE_CELL}", True),
    ("— Venue fee (10% revenue)", f"=-Assumptions!{VENUE_FEE_CELL}", False),
    ("— COGS (variable costs)", f"=-Assumptions!{COGS_CELL}", False),
]

r = 3
section_header(ws2, f"A{r}", "PER-CUP WATERFALL", "C")
ws2.row_dimensions[r].height = 20
r += 1

items = [
    ("Revenue per cup",             f"=Assumptions!{PRICE_CELL}",      False),
    ("  (–) Venue fee (10%)",        f"=-Assumptions!{VENUE_FEE_CELL}", False),
    ("  (–) Frozen fruit",           f"=-Assumptions!B{var_rows_start}", False),
    ("  (–) Cup + lid",              f"=-Assumptions!B{var_rows_start+1}", False),
    ("  (–) Label",                  f"=-Assumptions!B{var_rows_start+2}", False),
    ("  (–) Electricity",            f"=-Assumptions!B{var_rows_start+3}", False),
    ("  (–) Cleaning supplies",      f"=-Assumptions!B{var_rows_start+4}", False),
]

for lbl, formula, _ in items:
    label(ws2, f"A{r}", lbl)
    euro(ws2, f"B{r}", formula)
    ws2.row_dimensions[r].height = 16
    r += 1

r += 0  # no extra gap before margin
label(ws2, f"A{r}", "CONTRIBUTION MARGIN per cup", bold=True)
euro(ws2, f"B{r}", f"=Assumptions!{PRICE_CELL}-Assumptions!{VENUE_FEE_CELL}-Assumptions!{COGS_CELL}")
ws2[f"A{r}"].fill = fill(LIGHT_BG)
ws2[f"B{r}"].fill = fill(LIGHT_BG)
ws2[f"B{r}"].font = font(bold=True)
CONTRIB_CELL = f"B{r}"
ws2.row_dimensions[r].height = 18
r += 1

label(ws2, f"A{r}", "Contribution margin %", bold=False)
pct(ws2, f"B{r}", f"={CONTRIB_CELL}/Assumptions!{PRICE_CELL}")
ws2.row_dimensions[r].height = 16

# ── BREAK-EVEN ────────────────────────────────────────────────────────────────
r += 2
section_header(ws2, f"A{r}", "BREAK-EVEN ANALYSIS", "C")
ws2.row_dimensions[r].height = 20
r += 1

label(ws2, f"A{r}", "Fixed costs per month")
euro(ws2, f"B{r}", f"=Assumptions!{FIXED_CELL}")
ws2.row_dimensions[r].height = 16
r += 1

label(ws2, f"A{r}", "Break-even cups / month", bold=True)
ws2[f"B{r}"].value = f"=CEILING(Assumptions!{FIXED_CELL}/{CONTRIB_CELL},1)"
ws2[f"B{r}"].number_format = "#,##0"
ws2[f"B{r}"].font = font(bold=True)
ws2[f"A{r}"].fill = fill(LIGHT_BG)
ws2[f"B{r}"].fill = fill(LIGHT_BG)
BEP_CELL = f"B{r}"
ws2.row_dimensions[r].height = 18
r += 1

label(ws2, f"A{r}", "Break-even cups / day (30-day month)")
ws2[f"B{r}"].value = f"={BEP_CELL}/30"
ws2[f"B{r}"].number_format = "0.0"
ws2.row_dimensions[r].height = 16

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — MONTHLY SCENARIOS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Monthly Scenarios")
ws3.column_dimensions["A"].width = 30

scenarios = [50, 100, 150, 200, 300, 400]
for i, s in enumerate(scenarios):
    col = get_column_letter(i + 2)
    ws3.column_dimensions[col].width = 14

ws3.merge_cells(f"A1:{get_column_letter(len(scenarios)+1)}1")
ws3["A1"] = "Monthly Scenarios — Cups Sold per Month"
ws3["A1"].font = font(bold=True, size=14, color=WHITE)
ws3["A1"].fill = fill(DARK_BG)
ws3["A1"].alignment = align("center", "center")
ws3.row_dimensions[1].height = 28

# Column headers
ws3["A3"] = ""
for i, cups in enumerate(scenarios):
    col = get_column_letter(i + 2)
    ws3[f"{col}3"] = f"{cups} cups/mo"
    ws3[f"{col}3"].font = font(bold=True, color=WHITE)
    ws3[f"{col}3"].fill = fill(MID_BG)
    ws3[f"{col}3"].alignment = align("center")
    ws3.row_dimensions[3].height = 20

row_defs = [
    ("Cups / month",               "{cups}",          "#,##0",      False),
    ("Revenue",                    "{cups}*Assumptions!"+PRICE_CELL, "#,##0.00€", False),
    ("  Venue fee (10%)",          "-{cups}*Assumptions!"+VENUE_FEE_CELL, "#,##0.00€", False),
    ("  COGS",                     "-{cups}*Assumptions!"+COGS_CELL,      "#,##0.00€", False),
    ("Contribution",               "{cups}*'Unit Economics'!"+CONTRIB_CELL, "#,##0.00€", True),
    ("Fixed costs",                "-Assumptions!"+FIXED_CELL,             "#,##0.00€", False),
    ("NET PROFIT / month",         "{cups}*'Unit Economics'!"+CONTRIB_CELL+"-Assumptions!"+FIXED_CELL, "#,##0.00€", True),
    ("NET PROFIT / year",          "({cups}*'Unit Economics'!"+CONTRIB_CELL+"-Assumptions!"+FIXED_CELL+")*12", "#,##0.00€", True),
    ("Payback period (months)",    "=IF({cups}*'Unit Economics'!"+CONTRIB_CELL+">Assumptions!"+FIXED_CELL+",Assumptions!B9/{cups}/'Unit Economics'!"+CONTRIB_CELL+',"-")', "0.0", True),
]

# Build payback formula inline (simpler)
payback_template = (
    "=IF({cups}*'Unit Economics'!{contrib}>Assumptions!{fixed},"
    "Assumptions!B{total_capex_row}/({cups}*'Unit Economics'!{contrib}-Assumptions!{fixed})"
    ',"-")'
)

row_labels = [
    "Cups / month",
    "Revenue",
    "  (–) Venue fee (10%)",
    "  (–) COGS (total)",
    "Contribution margin",
    "  (–) Fixed costs",
    "NET PROFIT before tax / month",
    "  (–) IRC + derrama (22.5%)",
    "NET PROFIT after tax / month",
    "NET PROFIT after tax / year",
    "Payback — total CAPEX (months)",
]

highlight_rows = {4, 6, 8, 9, 10}   # 0-indexed from row_labels
r = 4
for li, lbl in enumerate(row_labels):
    bold = li in highlight_rows
    bg = LIGHT_BG if bold else WHITE
    label(ws3, f"A{r}", lbl, bold=bold)
    ws3[f"A{r}"].fill = fill(bg)
    ws3.row_dimensions[r].height = 16

    for i, cups in enumerate(scenarios):
        col = get_column_letter(i + 2)
        cell = ws3[f"{col}{r}"]

        if li == 0:
            cell.value = cups
            cell.number_format = "#,##0"
        elif li == 1:
            cell.value = f"={cups}*Assumptions!{PRICE_CELL}"
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 2:
            cell.value = f"=-{cups}*Assumptions!{VENUE_FEE_CELL}"
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 3:
            cell.value = f"=-{cups}*Assumptions!{COGS_CELL}"
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 4:
            cell.value = f"={cups}*'Unit Economics'!{CONTRIB_CELL}"
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 5:
            cell.value = f"=-Assumptions!{FIXED_CELL}"
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 6:
            # Net profit before tax
            cell.value = f"={cups}*'Unit Economics'!{CONTRIB_CELL}-Assumptions!{FIXED_CELL}"
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 7:
            # Tax = MAX(0, profit_before_tax) * effective_tax_rate
            cell.value = (
                f"=-MAX(0,{cups}*'Unit Economics'!{CONTRIB_CELL}-Assumptions!{FIXED_CELL})"
                f"*Assumptions!{EFFECTIVE_TAX_CELL}"
            )
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 8:
            # Net profit after tax / month
            cell.value = (
                f"=({cups}*'Unit Economics'!{CONTRIB_CELL}-Assumptions!{FIXED_CELL})"
                f"-MAX(0,{cups}*'Unit Economics'!{CONTRIB_CELL}-Assumptions!{FIXED_CELL})"
                f"*Assumptions!{EFFECTIVE_TAX_CELL}"
            )
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 9:
            # Net profit after tax / year
            cell.value = (
                f"=(({cups}*'Unit Economics'!{CONTRIB_CELL}-Assumptions!{FIXED_CELL})"
                f"-MAX(0,{cups}*'Unit Economics'!{CONTRIB_CELL}-Assumptions!{FIXED_CELL})"
                f"*Assumptions!{EFFECTIVE_TAX_CELL})*12"
            )
            cell.number_format = u'#,##0.00\u20ac'
        elif li == 10:
            cell.value = payback_template.format(
                cups=cups,
                contrib=CONTRIB_CELL,
                fixed=FIXED_CELL,
                total_capex_row=total_capex_row
            )
            cell.number_format = '0.0" mo"'

        cell.font = font(bold=bold)
        cell.fill = fill(bg)
        cell.alignment = align("right")
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
out_path = r"d:\Claude - PA\projects\smoothie-machine\financials\business-model.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
